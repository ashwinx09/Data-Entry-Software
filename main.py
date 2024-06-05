import tkinter as tk
from tkinter import *
from tkinter.ttk import Combobox
import pathlib
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk

# Main frame code
root = tk.Tk()
root.title("Data Entry Application")
root.geometry("800x500")
root.resizable(False, False)
root.configure(bg="ghost white")

# Backend excel file code
file_path = pathlib.Path("Backend_data.xlsx")
try:
    if file_path.exists():
        file = load_workbook(file_path)
        sheet = file.active
    else:
        file = Workbook()
        sheet = file.active
        sheet["A1"] = "Full Name"
        sheet["B1"] = "Phone Number"
        sheet["C1"] = "Age"
        sheet["D1"] = "Gender"
        sheet["E1"] = "Email"
        file.save(file_path)
except Exception as e:
    print("Error occurred while working with Excel file:", e)

# Submit function
def submit():
    global file, sheet
    name = nameValue.get()
    phone_number = contactValue.get()
    age = ageValue.get()
    gender = gender_combobox.get()
    email = emailValue.get()
    sheet.append([name, phone_number, age, gender, email])
    file.save(file_path)
    clear()

# Clear function
def clear():
    nameValue.set("")
    contactValue.set("")
    ageValue.set("")
    gender_combobox.current(0)
    emailValue.set("")

# Button color configurations
def on_enter(e):
    e.widget['background'] = 'firebrick4'  # When not hovered
def on_leave(e):
    e.widget['background'] = 'firebrick2'  # When hovered

# Application logo
image_path = pathlib.Path("logo.png")
try:
    if image_path.exists():
        image = Image.open(image_path)
        icon_image = ImageTk.PhotoImage(image)
        root.wm_iconphoto(False, icon_image)
    else:
        print("Image Error: logo.png not found")
except Exception as e:
    print("Error occurred while loading logo:", e)

# Main configurations
Label(root, text="DATA~ENTRY", font="Helvetica 25 bold", bg="ghost white", fg="firebrick2").place(x=80, y=20)
Label(root, text="NAME", font="Helvetica 18 bold", bg="ghost white", fg="#2F4F4F").place(x=80, y=94)
Label(root, text="CONTACT", font="Helvetica 18 bold", bg="ghost white", fg="#2F4F4F").place(x=40, y=145)
Label(root, text="AGE", font="Helvetica 18 bold", bg="ghost white", fg="#2F4F4F").place(x=80, y=194)
Label(root, text="GENDER", font="Helvetica 18 bold", bg="ghost white", fg="#2F4F4F").place(x=430, y=195)
Label(root, text="EMAIL", font="Helvetica 18 bold", bg="ghost white", fg="#2F4F4F").place(x=80, y=245)

nameValue = StringVar()
contactValue = StringVar()
ageValue = StringVar()
emailValue = StringVar()

# Entry box configurations
nameEntry = Entry(root, textvariable=nameValue, width=50, bd=2, font=19)
nameEntry.place(x=200, y=100)
contactEntry = Entry(root, textvariable=contactValue, width=50, bd=2, font=19)
contactEntry.place(x=200, y=150)
ageEntry = Entry(root, textvariable=ageValue, width=18, bd=2, font=19)
ageEntry.place(x=200, y=200)
gender_combobox = Combobox(root, values=["Male", "Female", "Other"], font="Helvetica 15", state="readonly", width=11)
gender_combobox.place(x=560, y=200)
gender_combobox.current(0)
emailEntry = Entry(root, textvariable=emailValue, width=50, bd=2, font=19)
emailEntry.place(x=200, y=250)

# Button configurations
submit_button = Button(root, text="SUBMIT", bg="firebrick2", fg="white", width=9, height=1, font="Helvetica 15 bold", command=submit)
submit_button.place(x=210, y=350)
clear_button = Button(root, text="CLEAR", bg="firebrick2", fg="white", width=9, height=1, font="Helvetica 15 bold", command=clear)
clear_button.place(x=355, y=350)
exit_button = Button(root, text="EXIT", bg="firebrick2", fg="white", width=9, height=1, font="Helvetica 15 bold", command=lambda: root.destroy())
exit_button.place(x=500, y=350)

# Button bindings
for button in [submit_button, clear_button, exit_button]:
    button.bind("<Enter>", on_enter)
    button.bind("<Leave>", on_leave)

root.mainloop()


